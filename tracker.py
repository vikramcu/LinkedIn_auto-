import os
from datetime import datetime
from openpyxl import Workbook, load_workbook

class TrackingManager:
    def __init__(self, excel_path="applications.xlsx"):
        self.excel_path = excel_path
        self._initialize_tracker()

    def _initialize_tracker(self):
        if not os.path.exists(self.excel_path):
            wb = Workbook()
            ws = wb.active
            ws.title = "Applications"
            # Define Headers
            ws.append(["Date", "Company", "Job Title", "Link", "Status"])
            wb.save(self.excel_path)

    def log_application(self, company, job_title, link, status):
        # 1. Log to Excel
        try:
            wb = load_workbook(self.excel_path)
            ws = wb.active
            
            date_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ws.append([date_str, company, job_title, link, status])
            
            wb.save(self.excel_path)
        except Exception as e:
            print(f"Failed to log to Excel: {e}")
